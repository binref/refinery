from __future__ import annotations

import enum
import functools
import io
import re

from datetime import datetime
from fnmatch import fnmatch
from typing import TYPE_CHECKING, Iterable

import defusedxml

from refinery.lib.structures import MemoryFile
from refinery.lib.tools import NoLogging
from refinery.lib.types import Param
from refinery.units import Arg, Unit

if TYPE_CHECKING:
    from openpyxl import Workbook as PyxlWorkbook
    from openpyxl.worksheet.worksheet import Worksheet as PyxlSheet
    from pyxlsb2 import Workbook as XlsbWorkbook
    from pyxlsb2.records import SheetRecord as XlsbSheet
    from xlrd2 import Book as XlrdWorkbook


defusedxml.defuse_stdlib()


def _ref2rc(ref: str):
    match = re.match(R'^([A-Z]+)(\d+)$', ref)
    if not match:
        raise ValueError
    col = functools.reduce(lambda acc, c: (acc * 26) + c, (ord(c) - 0x40 for c in match[1]), 0)
    row = int(match[2], 10)
    return row, col


def _rc2ref(row: int, col: int):
    if row <= 0:
        raise ValueError
    if col <= 0:
        raise ValueError
    alphabetic = ''
    while col:
        col, letter = divmod(col - 1, 26)
        alphabetic = chr(0x41 + letter) + alphabetic
    return F'{alphabetic}{row}'


class SheetReference:

    def _parse_sheet(self, token: str):
        try:
            sheet, token = token.rsplit('#', 1)
        except ValueError:
            sheet = None
        else:
            try:
                sheet = int(sheet, 0) - 1
            except (TypeError, ValueError):
                if sheet[0] in ('"', "'") and sheet[~0] == sheet[0] and len(sheet) > 2:
                    sheet = sheet[1:~1]
        return sheet, token

    def _parse_range(self, token: str):
        try:
            start, end = token.split(':')
            return start, end
        except ValueError:
            return token, token

    @staticmethod
    def _parse_token(token: str):
        try:
            row, col = _ref2rc(token)
        except ValueError:
            row, col = (int(x, 0) for x in token.split('.'))
        if row <= 0:
            raise ValueError(F'row must be positive, {row} is an invalid value')
        if col <= 0:
            raise ValueError(F'col must be positive, {col} is an invalid value')
        return row, col

    def __init__(self, sheet_reference=None):
        self.lbound = 1, 1
        self.ubound = None
        if sheet_reference is None:
            self.sheet = None
            return
        self.sheet, token = self._parse_sheet(sheet_reference)
        if not token:
            return
        try:
            start, stop = (self._parse_token(x) for x in self._parse_range(token))
        except Exception:
            self.sheet = sheet_reference
        else:
            row_min = min(start[0], stop[0])
            col_min = min(start[1], stop[1])
            row_max = max(start[0], stop[0])
            col_max = max(start[1], stop[1])
            self.lbound = (row_min, col_min)
            self.ubound = (row_max, col_max)

    def match(self, index: int, name: str):
        if self.sheet is None:
            return True
        if isinstance(self.sheet, int):
            return self.sheet == index
        return self.sheet == name or fnmatch(name, self.sheet)

    def cells(self, row_max, col_max):
        if self.ubound is not None:
            row_max, col_max = self.ubound
        row, col = self.lbound
        colstart = col
        while True:
            yield row, col
            if col < col_max:
                col += 1
            elif row < row_max:
                row, col = row + 1, colstart
            else:
                break

    def __contains__(self, ref):
        if self.ubound is None:
            return True
        if not isinstance(ref, tuple):
            ref = self._parse_token(ref)
        row, col = ref
        if row not in range(self.lbound[0], self.ubound[0] + 1):
            return False
        if col not in range(self.lbound[1], self.ubound[1] + 1):
            return False
        return True


class Workbook:

    workbook: XlsbWorkbook | XlrdWorkbook | PyxlWorkbook

    class _xlmode(enum.IntEnum):
        openpyxl = 1
        xlrd = 2
        pyxlsb2 = 3

    def __init__(self, data, unit: _ExcelUnit):
        def openpyxl():
            return unit._openpyxl.load_workbook(MemoryFile(data), read_only=True)

        def pyxlsb2():
            return unit._pyxlsb2.open_workbook(MemoryFile(data))

        def xlrd():
            verbose = max(unit.log_level.verbosity - 1, 0)
            log = unit._get_logger_io()
            return unit._xlrd.open_workbook(
                file_contents=data, logfile=log, verbosity=verbose, on_demand=True)

        exception = None

        for mode, loader in [
            (self._xlmode.openpyxl, openpyxl),
            (self._xlmode.xlrd, xlrd),
            (self._xlmode.pyxlsb2, pyxlsb2)
        ]:
            try:
                self.workbook = loader()
            except Exception as e:
                exception = e
            else:
                self.mode = mode
                exception = None
                break

        if exception:
            raise exception

    def sheets(self):
        if self.mode is self._xlmode.openpyxl:
            pyxl: PyxlWorkbook = self.workbook
            yield from pyxl.sheetnames
            return
        if self.mode is self._xlmode.xlrd:
            xlrd: XlrdWorkbook = self.workbook
            yield from xlrd.sheet_names()
            return
        if self.mode is self._xlmode.pyxlsb2:
            xlsb: XlsbWorkbook = self.workbook
            it: Iterable[XlsbSheet] = xlsb.sheets
            yield from (rec.name for rec in it)

    def get_sheet_data(self, name: str):
        def _sanitize(value):
            if value is None:
                return None
            if isinstance(value, str):
                return value
            try:
                it = iter(value)
            except Exception:
                pass
            else:
                return [_sanitize(v) for v in it]
            if isinstance(value, float):
                if float(int(value)) == value:
                    return int(value)
            if isinstance(value, datetime):
                return value.isoformat(' ', 'seconds')
            return str(value)

        def _padded(data: list[list[str]]):
            ncols = max((len(row) for row in data), default=0)
            for row in data:
                row.extend([None] * (ncols - len(row)))
            return data

        if self.mode is self._xlmode.openpyxl:
            pyxl_wbook: PyxlWorkbook = self.workbook
            pyxl_sheet: PyxlSheet = pyxl_wbook[name]
            with NoLogging():
                data = _padded(_sanitize(pyxl_sheet.iter_rows(values_only=True)))
        elif self.mode is self._xlmode.pyxlsb2:
            xlsb_wbook: XlsbWorkbook = self.workbook
            xlsb_sheet = xlsb_wbook.get_sheet_by_name(name)
            data = _padded(_sanitize(xlsb_sheet.rows()))
        elif self.mode is self._xlmode.xlrd:
            xlrd_wbook: XlrdWorkbook = self.workbook
            xlrd_sheet = xlrd_wbook.sheet_by_name(name)
            data = []
            for r in range(xlrd_sheet.nrows):
                row = []
                for c in range(xlrd_sheet.ncols):
                    try:
                        row.append(_sanitize(xlrd_sheet.cell_value(r, c)))
                    except IndexError:
                        row.append(None)
                data.append(row)
        else:
            raise RuntimeError(F'Invalid mode {self.mode!r}.')

        return data


class _ExcelUnit(Unit, abstract=True):

    @Unit.Requires('xlrd2', ['formats', 'office', 'extended'])
    def _xlrd():
        import xlrd2
        return xlrd2

    @Unit.Requires('openpyxl', ['formats', 'office', 'extended'])
    def _openpyxl():
        import openpyxl
        return openpyxl

    @Unit.Requires('pyxlsb2', ['formats', 'office', 'extended'])
    def _pyxlsb2():
        import pyxlsb2
        return pyxlsb2

    def _get_logger_io(self):
        class logger(io.TextIOBase):
            unit = self

            def write(self, string: str):
                string = string.strip()
                if not string or '\n' in string:
                    return
                if re.search(R'^[A-Z]+:', string) or '***' in string:
                    self.unit.log_debug(string)

        return logger()


class xlxtr(_ExcelUnit):
    """
    Extract data from Microsoft Excel documents, both Legacy and new XML type documents. A sheet
    reference is of the form `B1` or `1.2`, both specifying the first cell of the second column.
    A cell range can be specified as `B1:C12`, or `1.2:C12`, or `1.2:12.3`. Finally, the unit will
    always refer to the first sheet in the document and to change this, specify the sheet name or
    index separated by a hashtag, i.e. `sheet#B1:C12` or `1#B1:C12`. Note that indices are
    1-based. To get all elements of one sheet, use `sheet#`. The unit If parsing a sheet reference
    fails, the script will assume that the given reference specifies a sheet.
    """
    def __init__(
        self,
        *references: Param[SheetReference, Arg(
            metavar='reference',
            type=SheetReference,
            help=(
                'A sheet reference to be extracted. '
                'If no sheet references are given, the unit lists all sheet names.'
            )
        )]
    ):
        if not references:
            references = SheetReference('*'),
        super().__init__(references=references)

    def process(self, data):
        try:
            wb = Workbook(data, self)
        except ImportError:
            raise
        except Exception as E:
            raise ValueError('Input not recognized as Excel document.') from E
        for ref in self.args.references:
            ref: SheetReference
            for k, name in enumerate(wb.sheets()):
                if not ref.match(k, name):
                    continue
                try:
                    data = wb.get_sheet_data(name)
                except Exception as error:
                    self.log_info(F'error reading sheet {name}:', error)
                    continue
                for r, row in enumerate(data, 1):
                    for c, value in enumerate(row, 1):
                        if (r, c) not in ref:
                            continue
                        if value is None:
                            continue
                        yield self.labelled(
                            str(value).encode(self.codec),
                            row=r,
                            col=c,
                            ref=_rc2ref(r, c),
                            sheet=name
                        )
